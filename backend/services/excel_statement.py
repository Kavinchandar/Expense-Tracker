"""Extract transaction rows from XLS/XLSX statement exports."""

from __future__ import annotations

import io
import re
from datetime import date, datetime, timedelta
from decimal import Decimal
from typing import Any

from dateutil import parser as du_parser

from services.statement_parse_utils import finalize_parsed_rows


HEADER_ALIASES = {
    "date": (
        "date",
        "tran date",
        "txn date",
        "transaction date",
        "transaction dt",
        "transaction day",
        "value date",
        "posting date",
        "posted date",
    ),
    "description": (
        "description",
        "transaction description",
        "transaction remarks",
        "transaction narration",
        "transaction particulars",
        "transaction details",
        "narrative",
        "narration details",
        "narration",
        "particulars",
        "remarks",
        "details",
        "transaction details",
    ),
    "amount": ("amount", "transaction amount"),
    "debit": (
        "debit",
        "withdrawal",
        "withdrawal amt",
        "withdrawal amount",
        "debit amt",
        "debit amount",
        "dr",
        "debits",
    ),
    "credit": (
        "credit",
        "deposit",
        "deposit amt",
        "deposit amount",
        "credit amt",
        "credit amount",
        "cr",
        "credits",
    ),
    "balance": (
        "balance",
        "running balance",
        "closing balance",
        "available balance",
        "avail balance",
        "avail bal",
        "balance amount",
    ),
}


def _normalize_header(value: Any) -> str:
    s = " ".join(str(value or "").strip().lower().split())
    s = re.sub(r"[^a-z0-9 ]+", "", s)
    return s


def _to_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float, Decimal)):
        return float(value)
    s = str(value).strip()
    if not s:
        return None
    neg = False
    if s.startswith("(") and s.endswith(")"):
        neg = True
        s = s[1:-1]
    s_low = s.lower()
    if re.search(r"\s*dr\.?$", s_low):
        neg = True
        s = re.sub(r"\s*dr\.?$", "", s, flags=re.I).strip()
    if re.search(r"\s*cr\.?$", s_low):
        s = re.sub(r"\s*cr\.?$", "", s, flags=re.I).strip()
    if s.endswith("-"):
        neg = True
        s = s[:-1].strip()
    s = (
        s.replace(",", "")
        .replace("$", "")
        .replace("₹", "")
        .replace("€", "")
        .replace("\u20b9", "")
        .strip()
    )
    if not s:
        return None
    try:
        val = float(s)
    except ValueError:
        return None
    if neg:
        return -abs(val)
    return val


def _excel_serial_to_date(raw: float) -> date:
    epoch = datetime(1899, 12, 30)
    return (epoch + timedelta(days=float(raw))).date()


def _to_date(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and float(value) > 1000:
        try:
            return _excel_serial_to_date(float(value))
        except Exception:
            return None
    s = str(value).strip()
    if not s:
        return None
    try:
        return du_parser.parse(s, dayfirst=True, yearfirst=False).date()
    except (ValueError, TypeError, du_parser.ParserError):
        return None


def _match_alias(header: str, key: str) -> bool:
    for alias in HEADER_ALIASES[key]:
        if header == alias or header.startswith(alias + " "):
            return True
    return False


def _detect_columns(rows: list[list[Any]]) -> dict[str, int] | None:
    for ridx, row in enumerate(rows[:40]):
        header_map: dict[str, int] = {}
        for cidx, value in enumerate(row):
            normalized = _normalize_header(value)
            if not normalized:
                continue
            for key in ("date", "description", "amount", "debit", "credit", "balance"):
                if key in header_map:
                    continue
                if _match_alias(normalized, key):
                    header_map[key] = cidx
                    break
        has_amount_shape = "amount" in header_map or (
            "debit" in header_map and "credit" in header_map
        )
        if "date" in header_map and has_amount_shape:
            if "description" not in header_map:
                guessed = _guess_description_column(
                    rows, start_row=ridx + 1, excluded_cols=set(header_map.values())
                )
                if guessed is not None:
                    header_map["description"] = guessed
        if "date" in header_map and "description" in header_map and has_amount_shape:
            header_map["__header_row__"] = ridx
            return header_map
    return None


def _guess_description_column(
    rows: list[list[Any]], start_row: int, excluded_cols: set[int]
) -> int | None:
    scores: dict[int, int] = {}
    for row in rows[start_row : start_row + 80]:
        for cidx, value in enumerate(row):
            if cidx in excluded_cols:
                continue
            if value is None:
                continue
            txt = str(value).strip()
            if len(txt) < 3:
                continue
            if _to_date(value) is not None:
                continue
            if _to_float(value) is not None:
                continue
            scores[cidx] = scores.get(cidx, 0) + 1
    if not scores:
        return None
    best_col, best_score = max(scores.items(), key=lambda x: x[1])
    if best_score < 2:
        return None
    return best_col


def _resolve_amount(row: list[Any], cols: dict[str, int]) -> float | None:
    if "amount" in cols:
        amount = _to_float(row[cols["amount"]] if cols["amount"] < len(row) else None)
        if amount is not None:
            return amount
    debit = _to_float(row[cols["debit"]] if cols.get("debit", -1) < len(row) else None)
    credit = _to_float(row[cols["credit"]] if cols.get("credit", -1) < len(row) else None)
    if debit is None and credit is None:
        return None
    return round((credit or 0.0) - abs(debit or 0.0), 2)


def _extract_rows_from_grid(grid_rows: list[list[Any]], source_tag: str) -> list[dict[str, Any]]:
    cols = _detect_columns(grid_rows)
    if not cols:
        return []
    out: list[dict[str, Any]] = []
    start = cols["__header_row__"] + 1
    date_col = cols["date"]
    desc_col = cols["description"]
    bal_col = cols.get("balance")
    for ridx, row in enumerate(grid_rows[start:], start=start + 1):
        if not row or all((str(v).strip() == "" for v in row if v is not None)):
            continue
        d = _to_date(row[date_col] if date_col < len(row) else None)
        if d is None:
            continue
        desc = str(row[desc_col] if desc_col < len(row) else "").strip()
        if len(desc) < 2:
            continue
        amt = _resolve_amount(row, cols)
        if amt is None or abs(amt) < 0.005:
            continue
        bal = _to_float(row[bal_col] if bal_col is not None and bal_col < len(row) else None)
        tid = f"{source_tag}:{d.isoformat()}|{amt:.2f}|{desc[:180]}|r{ridx}"
        parsed: dict[str, Any] = {
            "date": d,
            "description": desc[:1024],
            "amount": round(float(amt), 2),
            "transaction_id": tid,
        }
        if bal is not None:
            parsed["balance_after"] = round(float(bal), 2)
        out.append(parsed)
    return finalize_parsed_rows(out)


def extract_transaction_lines_from_xlsx(file_bytes: bytes) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except Exception as e:  # pragma: no cover - dependency/import error path
        raise RuntimeError("openpyxl is required for .xlsx support") from e
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    try:
        for sheet in wb.worksheets:
            grid_rows = [list(r) for r in sheet.iter_rows(values_only=True)]
            rows = _extract_rows_from_grid(grid_rows, source_tag=f"xlsx:{sheet.title}")
            if rows:
                return rows
    finally:
        wb.close()
    return []


def extract_transaction_lines_from_xls(file_bytes: bytes) -> list[dict[str, Any]]:
    try:
        import xlrd
    except Exception as e:  # pragma: no cover - dependency/import error path
        raise RuntimeError("xlrd is required for .xls support") from e
    book = xlrd.open_workbook(file_contents=file_bytes)
    for sheet in book.sheets():
        grid_rows = [sheet.row_values(ridx) for ridx in range(sheet.nrows)]
        rows = _extract_rows_from_grid(grid_rows, source_tag=f"xls:{sheet.name}")
        if rows:
            return rows
    return []
